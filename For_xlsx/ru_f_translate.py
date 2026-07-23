#!/usr/bin/env python3
"""
ru_f_translate.py — Конвертер текста в RU_F кодировку для .xlsx файлов.

Берёт читаемый русский текст из столбца TL, конвертирует кириллицу → байты
специализированного шрифта (cp1252-based) и записывает результат в столбец TLC.
Столбцы TL и Edit не изменяются (остаются читаемой кириллицей).
Результат сохраняется в новый файл с суффиксом _ruf.xlsx.

Использование:
    python ru_f_translate.py input.xlsx
    python ru_f_translate.py input.xlsx -o output.xlsx
    python ru_f_translate.py input.xlsx --preview        # показать первые N строк без сохранения
"""

import argparse
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Установите openpyxl:  pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# RU_F таблица трансляции (из bgi_tool.py)
# ─────────────────────────────────────────────────────────────────────────────

_RU_F_UPPER: dict[str, str] = {
    'A': 'А', 'B': 'Б', 'C': 'В', 'D': 'Г', 'E': 'Д', 'F': 'Е',
    'G': 'Ж', 'H': 'З', 'I': 'И', 'J': 'Й', 'K': 'К', 'L': 'Л',
    'M': 'М', 'N': 'Н', 'O': 'О', 'P': 'П', 'Q': 'Р', 'R': 'С',
    'S': 'Т', 'T': 'У', 'U': 'Ф', 'V': 'Х', 'W': 'Ц', 'X': 'Ч',
    'Y': 'Ш', 'Z': 'Щ',
}
_RU_F_LOWER: dict[str, str] = {k.lower(): v.lower() for k, v in _RU_F_UPPER.items()}

# Обратная таблица: кириллица → байт(ы) cp1252
_CYR_TO_LATIN: dict[str, str] = {}
for _lat, _cyr in _RU_F_UPPER.items():
    _CYR_TO_LATIN[_cyr] = _lat          # заглавные
for _lat, _cyr in _RU_F_LOWER.items():
    _CYR_TO_LATIN[_cyr] = _lat          # строчные

# Спецсимволы шрифта (символы без прямого латинского соответствия)
_CYR_SPECIAL: dict[str, str] = {
    'Ъ': '[',   'Ь': ']',   'ё': '`',
    'э': '{',   'ы': '|',   'я': '}',
    'Ы': '\xa1',
    '-': '#',
    'ь': '&',
    'ъ': '+',
    'Ю': '-',
    'ю': '$',
    '—': '#',
    'Я': '>',
    'Ё': '<',
    '«': '"',
    '»': '"',
    'Э': '=',
    'Й': 'J',   'й': 'j',
}

# Объединяем: сначала спецсимволы (приоритет), потом основные
_FULL_TABLE: dict[str, str] = {}
_FULL_TABLE.update(_CYR_TO_LATIN)
_FULL_TABLE.update(_CYR_SPECIAL)   # спецсимволы перекрывают (например, Й → J)


def apply_ru_f(text: str) -> str:
    """
    Применяет трансляцию RU_F к строке.
    Кириллица → соответствующий латинский/спецсимвол cp1252.
    Нетронутые символы (латиница, пунктуация и т.д.) остаются как есть.
    """
    if not text:
        return text
    out = []
    for ch in text:
        out.append(_FULL_TABLE.get(ch, ch))
    return ''.join(out)


# ─────────────────────────────────────────────────────────────────────────────
# Определение структуры листа
# ─────────────────────────────────────────────────────────────────────────────

TARGET_COLUMNS = {'TL', 'TLC', 'Edit'}

# Листы, которые нужно полностью пропускать — не читать и не изменять никак.
SKIP_SHEETS = {'hayami_aa13', 're_otoha_01c'}


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    """
    Ищет строку заголовков и возвращает (номер строки, {имя_колонки: индекс_0based}).
    Возвращает (None, {}) если заголовки не найдены.
    """
    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        col_map = {}
        for col_idx, val in enumerate(row):
            if val and str(val).strip() in TARGET_COLUMNS:
                col_map[str(val).strip()] = col_idx
        if col_map:
            return row_idx, col_map
    return None, {}


# ─────────────────────────────────────────────────────────────────────────────
# Основная обработка
# ─────────────────────────────────────────────────────────────────────────────

def translate_xlsx(input_path: str, output_path: str, preview: int = 0) -> None:
    print(f"Читаем: {input_path}")
    wb = openpyxl.load_workbook(input_path)

    total_converted = 0
    total_cells = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"  [{sheet_name}] — в списке SKIP_SHEETS, пропускаем полностью (не трогаем)")
            continue

        ws = wb[sheet_name]
        header_row_idx, col_map = find_header_row(ws)

        if 'TL' not in col_map or 'TLC' not in col_map:
            print(f"  [{sheet_name}] — не найден столбец TL и/или TLC, пропускаем")
            continue

        tl_idx = col_map['TL']
        tlc_idx = col_map['TLC']
        print(f"  [{sheet_name}] — источник: TL(col {tl_idx+1}), запись: TLC(col {tlc_idx+1})")

        sheet_converted = 0
        sheet_cells = 0
        preview_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1), start=header_row_idx + 1):
            if tl_idx >= len(row):
                continue
            src_cell = row[tl_idx]
            val = src_cell.value
            if val is None or str(val).strip() == '':
                continue

            original = str(val)
            translated = apply_ru_f(original)
            sheet_cells += 1

            if translated != original:
                sheet_converted += 1
                if preview > 0 and preview_count < preview:
                    print(f"    [TL→TLC] {original!r} → {translated!r}")
                    preview_count += 1

            if preview == 0:
                if tlc_idx >= len(row):
                    # строка короче, чем нужно для TLC — расширяем через прямую запись в лист
                    tlc_cell = ws.cell(row=row_idx, column=tlc_idx + 1)
                else:
                    tlc_cell = row[tlc_idx]
                tlc_cell.value = translated
                # openpyxl автоматически считает строки, начинающиеся с '=',
                # формулами (data_type='f'). После RU_F-перевода такие строки
                # часто появляются (например, 'Э' -> '='), и Excel потом
                # ломает файл, вырезая "битые формулы". Принудительно
                # возвращаем тип "текст".
                if translated.startswith('='):
                    tlc_cell.data_type = 's'

        print(f"    → ячеек с текстом: {sheet_cells}, изменено: {sheet_converted}")
        total_converted += sheet_converted
        total_cells += sheet_cells

    print(f"\nИтого: {total_cells} ячеек, {total_converted} изменено.")

    if preview > 0:
        print("\n[Режим --preview: файл не сохранён]")
        return

    wb.save(output_path)
    print(f"Сохранено → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Берёт текст из TL, конвертирует в кодировку RU_F и пишет в TLC (TL и Edit не трогает).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры:
  python ru_f_translate.py strings.xlsx
  python ru_f_translate.py strings.xlsx -o strings_ruf.xlsx
  python ru_f_translate.py strings.xlsx --preview 10
""",
    )
    parser.add_argument("input", help="Входной .xlsx файл")
    parser.add_argument("-o", "--output", default=None,
                        help="Выходной .xlsx файл (по умолчанию: input_ruf.xlsx)")
    parser.add_argument("--preview", type=int, default=0, metavar="N",
                        help="Показать N примеров конвертации без сохранения файла")

    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"Ошибка: файл не найден: {args.input}")
        sys.exit(1)

    if args.output is None:
        base, ext = os.path.splitext(args.input)
        args.output = base + "_ruf" + (ext if ext else ".xlsx")

    translate_xlsx(args.input, args.output, preview=args.preview)


if __name__ == "__main__":
    main()